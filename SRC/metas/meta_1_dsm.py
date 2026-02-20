import sys
import os
import csv
import openpyxl

# Add project root to path to import modules
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(project_root, 'SRC'))

from modules.dataloaders import scan_rem_files, get_rem_value
from modules.utils import normalize_path
from config import DIR_SERIE_A_ACTUAL, DIR_SERIE_A_ANTERIOR, AGNO_ACTUAL, AGNO_ANTERIOR

def calcular_meta_1():
    print("=== Calculando Meta 1: Recuperación del Desarrollo Psicomotor ===")
    
    # 1. Configuración
    TARGET_SHEET = "A03"
    # Columnas de 12 a 23 meses
    COLS = ['J', 'K', 'L', 'M']
    # Filas para denominador (Primera Evaluación - Riesgo)
    ROWS_DEN = [23]
    # Filas para numerador (Reevaluación: Normal y Normal con rezago)
    ROWS_NUM = [26, 28]
    
    # 2. Cargar todos los REM Serie A disponibles (actual y anterior)
    mapping_actual = scan_rem_files(DIR_SERIE_A_ACTUAL)
    mapping_anterior = scan_rem_files(DIR_SERIE_A_ANTERIOR)
    mapping = mapping_actual + mapping_anterior
    print(f"Se encontraron {len(mapping)} archivos REM en total.")

    # 3. Filtrar archivos para numerador y denominador según lógica de negocio
    numerador_files = [entry for entry in mapping if entry['year'] == AGNO_ACTUAL and 1 <= entry['month'] <= 12]
    denominador_files = [entry for entry in mapping if (
        (entry['year'] == AGNO_ANTERIOR and entry['month'] >= 10) or
        (entry['year'] == AGNO_ACTUAL and entry['month'] <= 9)
    )]
    print(f"Archivos para numerador: {[f['filename'] for f in numerador_files]}")
    print(f"Archivos para denominador: {[f['filename'] for f in denominador_files]}")

    # Estructura para acumular por centro
    centros = {}

    # Procesar Numerador
    for entry in numerador_files:
        code = entry['code']
        file_path = entry['path']
        print(f"Procesando numerador: {file_path} (Centro: {code})")
        if code not in centros:
            centros[code] = {'num': 0, 'den': 0}
        if not os.path.exists(file_path):
            print(f"Archivo no existe: {file_path}")
            continue
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if TARGET_SHEET in wb.sheetnames:
                sheet = wb[TARGET_SHEET]
                for col in COLS:
                    for row in ROWS_NUM:
                        cell = f"{col}{row}"
                        val = sheet[cell].value
                        print(f"Numerador {cell}: {val}")
                        if val and isinstance(val, (int, float)):
                            centros[code]['num'] += val
            else:
                print(f"Hoja {TARGET_SHEET} no encontrada en {file_path}")
            wb.close()
        except Exception as e:
            print(f"Error procesando numerador {entry['filename']}: {e}")

    # Procesar Denominador
    for entry in denominador_files:
        code = entry['code']
        file_path = entry['path']
        print(f"Procesando denominador: {file_path} (Centro: {code})")
        if code not in centros:
            centros[code] = {'num': 0, 'den': 0}
        if not os.path.exists(file_path):
            print(f"Archivo no existe: {file_path}")
            continue
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if TARGET_SHEET in wb.sheetnames:
                sheet = wb[TARGET_SHEET]
                for col in COLS:
                    for row in ROWS_DEN:
                        cell = f"{col}{row}"
                        val = sheet[cell].value
                        print(f"Denominador {cell}: {val}")
                        if val and isinstance(val, (int, float)):
                            centros[code]['den'] += val
            else:
                print(f"Hoja {TARGET_SHEET} no encontrada en {file_path}")
            wb.close()
        except Exception as e:
            print(f"Error procesando denominador {entry['filename']}: {e}")

    reporte = []
    
    # Estructura para acumular por centro
    # centros[code] = {'num': 0, 'den': 0}
    centros = {}

    for entry in mapping:
        code = entry['code']
        file_path = entry['path']
        filename = entry['filename']
        year = entry['year']
        month = entry['month']
        
        if code not in centros:
            centros[code] = {'num': 0, 'den': 0}
            
        if not os.path.exists(file_path):
            continue
            
        # Lógica de Fechas (Lag)
        # Numerador: Enero a Diciembre 2026
        is_numerator_period = (year == 2026)
        
        # Denominador: Octubre 2025 a Septiembre 2026
        is_denominator_period = False
        if year == 2025 and month and month >= 10:
            is_denominator_period = True
        elif year == 2026 and month and month <= 9:
            is_denominator_period = True
            
        if not is_numerator_period and not is_denominator_period:
            continue

        num_local = 0
        den_local = 0
        
        try:
             wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
             if TARGET_SHEET in wb.sheetnames:
                 sheet = wb[TARGET_SHEET]
                 
                 # Si corresponde al periodo del Numerador
                 if is_numerator_period:
                     for col in COLS:
                         cell = f"{col}{ROW_NUM}"
                         val = sheet[cell].value
                         if val and isinstance(val, (int, float)):
                             num_local += val
                             
                 # Si corresponde al periodo del Denominador
                 if is_denominator_period:
                     for col in COLS:
                         cell = f"{col}{ROW_DEN}"
                         val = sheet[cell].value
                         if val and isinstance(val, (int, float)):
                             den_local += val
                         
             wb.close()
        except Exception as e:
             print(f"Error procesando {filename}: {e}")
             continue
            
        centros[code]['num'] += num_local
        centros[code]['den'] += den_local
        
        # Debug info
        # print(f"Procesado {filename} (Mes:{month} Año:{year}): Num={num_local if is_numerator_period else 0}, Den={den_local if is_denominator_period else 0}")
    
    # Generar reporte final
    total_num = 0
    total_den = 0
    
    for code, data in centros.items():
        num = data['num']
        den = data['den']
        total_num += num
        total_den += den
        
        cumplimiento = (num / den * 100) if den > 0 else 0
        
        reporte.append({
            'Centro': code,
            'Meta_ID': 'Meta 1',
            'Indicador': 'DSM',
            'Numerador': num,
            'Denominador': den,
            'Cumplimiento': cumplimiento,
            'Meta_Fijada': 90.0,
            'Meta_Nacional': 90.0
        })

    # 4. Resultado Final Global
    cumplimiento_global = (total_num / total_den * 100) if total_den > 0 else 0
    
    print("\n=== RESULTADOS GLOBALES META 1 ===")
    print(f"Numerador Total (Recuperados): {total_num}")
    print(f"Denominador Total (Riesgo): {total_den}")
    print(f"Cumplimiento Actual: {cumplimiento_global:.2f}%")
    print(f"Meta Fijada: 90.0%")
    
    # Guardar reporte
    output_path = normalize_path(r"DATOS\reporte_meta_1_preliminar.csv")
    
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Centro', 'Meta_ID', 'Indicador', 'Numerador', 'Denominador', 'Cumplimiento', 'Meta_Fijada', 'Meta_Nacional'])
            writer.writeheader()
            writer.writerows(reporte)
        print(f"Reporte detallado guardado en: {output_path}")
    except Exception as e:
        print(f"Error escribiendo reporte: {e}")

if __name__ == "__main__":
    calcular_meta_1()
