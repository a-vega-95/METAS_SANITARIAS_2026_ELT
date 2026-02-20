import sys
import os
import csv
import openpyxl

# Add project root to path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(project_root, 'SRC'))

from modules.dataloaders import scan_rem_files, load_piv_data
from modules.utils import normalize_path
from config import DIR_SERIE_A_ACTUAL, PIV_FILE

def calcular_meta_3():
    print("=== Calculando Meta 3: Salud Bucal ===")
    
    # 1. Configuración
    DATA_DIR_A = DIR_SERIE_A_ACTUAL
    
    # Meta 3A: CERO (0-9 años)
    # Num: REM A03, Sección D.7. "Pauta CERO" -> Fila "TOTAL" -> Suma Col 5 a 24 (0 a 9 años)
    SHEET_3A = "A03"
    COLS_IDX_3A = range(5, 25) # 5 to 24 inclusive (<1 to 9 years, M+F)
    
    # Meta 3B: Libre de Caries (6 años)
    # Num: REM A09, Sección C. S48 + T48
    SHEET_3B = "A09"
    CELLS_3B = ["S48", "T48"]
    
    # Buscar archivo PIV más reciente
    piv_dir = normalize_path("DATOS/PIV")
    piv_files = [f for f in os.listdir(piv_dir) if f.startswith("PIV_") and f.endswith(".parquet")]
    if not piv_files:
        print(f"ERROR: No se encontró ningún archivo PIV válido en: {piv_dir}")
        return
    piv_files.sort(reverse=True)
    piv_file = os.path.join(piv_dir, piv_files[0])
    print(f"Usando archivo PIV: {piv_file}")

    # Validar encabezados del parquet
    import pyarrow.parquet as pq
    try:
        table = pq.read_table(piv_file)
        expected_cols = {'COD_CENTRO', 'EDAD_EN_FECHA_CORTE', 'ACEPTADO_RECHAZADO', 'GENERO', 'GENERO_NORMALIZADO'}
        parquet_cols = set(table.schema.names)
        if not expected_cols.issubset(parquet_cols):
            print(f"ERROR: El archivo PIV no es compatible por encabezados. Esperado: {expected_cols}, encontrado: {parquet_cols}")
            return
        piv_data = table.to_pylist()
    except Exception as e:
        print(f"ERROR al leer el archivo PIV: {e}")
        return

    mapping_a = scan_rem_files(DATA_DIR_A)
    print(f"Archivos REM A para meta 3: {[f['filename'] for f in mapping_a]}")

    for entry in mapping_a:
        code = entry['code']
        real_code = code
        if code[-1].isalpha() and code[:-1].isdigit():
             real_code = code[:-1}
        file_path = entry['path']
        year = entry['year']
        print(f"Procesando REM A: {file_path} (Centro: {real_code})")
        if year != 2026: continue
        if real_code not in num_3a:
            num_3a[real_code] = 0
            num_3b[real_code] = 0
        if not os.path.exists(file_path):
            print(f"Archivo no existe: {file_path}")
            continue
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            # Meta 3A (A03)
            if SHEET_3A in wb.sheetnames:
                ws = wb[SHEET_3A]
                target_row = None
                found_section = False
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=300, values_only=True), 1):
                    content = " ".join([str(c) for c in row[:5] if c])
                    if "PAUTA CERO" in content:
                        found_section = True
                        continue
                    if found_section and "TOTAL" in content:
                        target_row = row
                        break
                if target_row:
                    val_3a = 0
                    for idx in COLS_IDX_3A:
                        if idx < len(target_row):
                            v = target_row[idx]
                            print(f"Meta 3A columna {idx}: {v}")
                            if v and isinstance(v, (int, float)):
                                val_3a += v
                    num_3a[real_code] += val_3a
                else:
                    print(f"No se encontró fila TOTAL en sección PAUTA CERO en {file_path}")
            else:
                print(f"Hoja {SHEET_3A} no encontrada en {file_path}")
            # Meta 3B (A09)
            if SHEET_3B in wb.sheetnames:
                ws = wb[SHEET_3B]
                val_3b = 0
                for cell in CELLS_3B:
                    v = ws[cell].value
                    print(f"Meta 3B celda {cell}: {v}")
                    if v and isinstance(v, (int, float)):
                        val_3b += v
                num_3b[real_code] += val_3b
            else:
                print(f"Hoja {SHEET_3B} no encontrada en {file_path}")
            wb.close()
        except Exception as e:
            print(f"Error procesando {file_path}: {e}")
                     
                     if found_section and "TOTAL" in content:
                         target_row = row
                         break
                     
                     # Safety break/Reset if we go too far (e.g. next section)
                     # Assuming TOTAL is close (within 10 rows)
                 
                 if target_row:
                     val_3a = 0
                     for idx in COLS_IDX_3A:
                         if idx < len(target_row):
                             v = target_row[idx]
                             if v and isinstance(v, (int, float)):
                                 val_3a += v
                     num_3a[real_code] += val_3a
                     
             # Meta 3B (A09)
             if SHEET_3B in wb.sheetnames:
                 ws = wb[SHEET_3B]
                 val_3b = 0
                 for cell in CELLS_3B:
                     v = ws[cell].value
                     if v and isinstance(v, (int, float)):
                         val_3b += v
                 num_3b[real_code] += val_3b
                 
             wb.close()
        except: pass

    # Reporte
    all_centers = set(den_3a.keys()) | set(num_3a.keys())
    reporte = []
    
    for c in all_centers:
        # 3A
        n3a = num_3a.get(c, 0)
        d3a = den_3a.get(c, 0)
        c3a = (n3a/d3a*100) if d3a > 0 else 0
        reporte.append({
            'Centro': c, 'Meta_ID': 'Meta 3A', 'Indicador': 'CERO (0-9)',
            'Numerador': n3a, 'Denominador': d3a, 'Cumplimiento': c3a,
            'Meta_Fijada': 0.0, 'Meta_Nacional': 0.0 # TBD
        })
        
        # 3B
        n3b = num_3b.get(c, 0)
        d3b = den_3b.get(c, 0)
        c3b = (n3b/d3b*100) if d3b > 0 else 0
        reporte.append({
            'Centro': c, 'Meta_ID': 'Meta 3B', 'Indicador': 'Libre Caries (6)',
            'Numerador': n3b, 'Denominador': d3b, 'Cumplimiento': c3b,
            'Meta_Fijada': 21.0, 'Meta_Nacional': 21.0
        })
        
    output_path = normalize_path(r"DATOS\reporte_meta_3_preliminar.csv")
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Centro', 'Meta_ID', 'Indicador', 'Numerador', 'Denominador', 'Cumplimiento', 'Meta_Fijada', 'Meta_Nacional'])
            writer.writeheader()
            for r in reporte:
                writer.writerow(r)
        print(f"Reporte guardado en {output_path}")
    except: pass

if __name__ == "__main__":
    calcular_meta_3()
