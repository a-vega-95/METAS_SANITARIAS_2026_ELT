import sys
import os
import csv
import openpyxl
import pyarrow.parquet as pq

# Add project root to path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(project_root, 'SRC'))

from modules.dataloaders import scan_rem_files, load_piv_data
from modules.utils import normalize_path
from config import DIR_SERIE_P_ACTUAL, PIV_FILE

def calcular_meta_4():
    print("=== Calculando Meta 4: Diabetes Mellitus Tipo 2 (DM2) ===")
    
    # Configuración
    DATA_DIR = DIR_SERIE_P_ACTUAL
    
    # 4A: Cobertura Efectiva
    # Num: REM P04, Sección B. C36 + C37 (Compensados)
    # Den: Personas 15+ con DM2 Estimadas (Prev 12.3%)
    PREVALENCIA_DM2 = 0.123 
    
    SHEET = "P4"
    CELLS_4A_NUM = ["C36", "C37"]
    
    # 4B: Pie Diabético
    # Num: Evaluacion Pie Vigente (C61+C62+C63+C64)
    # Den: DM2 Bajo Control (C17)
    CELLS_4B_NUM = ["C61", "C62", "C63", "C64"]
    CELLS_4B_DEN = ["C17"]
    
    # Buscar archivo PIV más reciente
    piv_dir = normalize_path("DATOS/PIV")
    piv_files = [f for f in os.listdir(piv_dir) if f.startswith("PIV_") and f.endswith(".parquet")]
    if not piv_files:
        print(f"ERROR: No se encontró ningún archivo PIV válido en: {piv_dir}")
        return
    piv_files.sort(reverse=True)
    piv_file = os.path.join(piv_dir, piv_files[0])
    print(f"Usando archivo PIV: {piv_file}")

    # Validar encabezados del parquet
    import pyarrow.parquet as pq
    try:
        table = pq.read_table(piv_file)
        expected_cols = {'COD_CENTRO', 'EDAD_EN_FECHA_CORTE', 'ACEPTADO_RECHAZADO', 'GENERO', 'GENERO_NORMALIZADO'}
        parquet_cols = set(table.schema.names)
        if not expected_cols.issubset(parquet_cols):
            print(f"ERROR: El archivo PIV no es compatible por encabezados. Esperado: {expected_cols}, encontrado: {parquet_cols}")
            return
        piv_data = table.to_pylist()
    except Exception as e:
        print(f"ERROR al leer el archivo PIV: {e}")
        return

    mapping = scan_rem_files(DATA_DIR)

    # 1. Denominadores 4A (Estimados)
    poblacion_15_mas = {}
    
    for row in piv_data:
        centro = row.get('COD_CENTRO', '')
        edad = row.get('EDAD_EN_FECHA_CORTE')
        if edad is None: edad = -1
        estado = row.get('ACEPTADO_RECHAZADO', '')
        
        if estado == 'ACEPTADO' and edad >= 15:
            if centro not in poblacion_15_mas:
                 poblacion_15_mas[centro] = 0
            poblacion_15_mas[centro] += 1
            
    denominadores_4a = {k: round(v * PREVALENCIA_DM2) for k, v in poblacion_15_mas.items()}
    
    # 2. Numeradores y Denominadores REM
    numeradores_4a = {}
    numeradores_4b = {}
    denominadores_4b = {}
    
    for entry in mapping:
        raw_code = entry['code']
        real_code = raw_code
        if raw_code[-1].isalpha() and raw_code[:-1].isdigit():
             real_code = raw_code[:-1]
             
        file_path = entry['path']
        
        if real_code not in numeradores_4a:
            numeradores_4a[real_code] = 0
            numeradores_4b[real_code] = 0
            denominadores_4b[real_code] = 0
            
        if not os.path.exists(file_path): continue
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if SHEET in wb.sheetnames:
                sheet = wb[SHEET]
                
                # 4A Num (Compensados)
                # Search for rows containing "HbA1C"
                # Logic: C36 + C37 in user request -> Corresponds to <7% and <8%
                # Dump showed them at Rows 30 and 31.
                
                rows_4a = []
                for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
                    row_str = " ".join([str(c) for c in row[:5] if c])
                    if "HbA1C<7%" in row_str or "HbA1C<8%" in row_str:
                        # Value is usually in Column C (Index 2)
                        # Check if it has a value
                        if len(row) > 2 and isinstance(row[2], (int, float)):
                            numeradores_4a[real_code] += row[2]

                # 4B Num (Pie Vigente)
                # C61+C62+C63+C64 in user request.
                # Need to find "evaluación vigente del pie"
                # Dump Row 61: "Con evaluación vigente del pie..." -> Riesgo bajo
                # Row 62: Riesgo moderado
                # Row 63: Riesgo alto
                # Row 64: Riesgo máximo
                # So we sum the 4 rows starting from "evaluación vigente del pie"
                
                found_pie = False
                pie_rows_count = 0
                for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
                    row_str = " ".join([str(c) for c in row[:5] if c])
                    if "evaluación vigente del pie" in row_str:
                         found_pie = True
                    
                    if found_pie and pie_rows_count < 4:
                         if len(row) > 2 and isinstance(row[2], (int, float)):
                             numeradores_4b[real_code] += row[2]
                         pie_rows_count += 1
                         
                # 4B Den (Bajo Control)
                # C17 from user.
                # Dump Row 17: "Diabetes Mellitus tipo 2" in Section A (Row 17)
                # Value at Col C (Index 2): 1300.
                # Dynamic search: "Diabetes Mellitus tipo 2" in Section A.
                # Section A starts around Row 8.
                for row in sheet.iter_rows(min_row=10, max_row=25, values_only=True): # Narrow range for Sec A
                    row_str = " ".join([str(c) for c in row[:5] if c])
                    if "Diabetes Mellitus tipo 2" in row_str:
                         if len(row) > 2 and isinstance(row[2], (int, float)):
                             denominadores_4b[real_code] += row[2]
                             break # Only one row in Section A
            
            wb.close()
        except:
            pass

            
    # Reporte
    all_centers = set(denominadores_4a.keys()) | set(numeradores_4a.keys())
    reporte = []
    
    for c in all_centers:
        # Meta 4A
        num_4a = numeradores_4a.get(c, 0)
        den_4a = denominadores_4a.get(c, 0)
        cump_4a = (num_4a/den_4a*100) if den_4a > 0 else 0
        
        reporte.append({
            'Centro': c, 'Meta_ID': 'Meta 4A', 'Indicador': 'Compensación DM2',
            'Numerador': num_4a, 'Denominador': den_4a, 'Cumplimiento': cump_4a,
            'Meta_Fijada': 29.0, 'Meta_Nacional': 29.0
        })
        
        # Meta 4B
        num_4b = numeradores_4b.get(c, 0)
        den_4b = denominadores_4b.get(c, 0)
        cump_4b = (num_4b/den_4b*100) if den_4b > 0 else 0
        
        reporte.append({
            'Centro': c, 'Meta_ID': 'Meta 4B', 'Indicador': 'Pie Diabético',
            'Numerador': num_4b, 'Denominador': den_4b, 'Cumplimiento': cump_4b,
            'Meta_Fijada': 90.0, 'Meta_Nacional': 90.0
        })
    
    # Output
    output_path = normalize_path(r"DATOS\reporte_meta_4a_preliminar.csv")
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Centro', 'Meta_ID', 'Indicador', 'Numerador', 'Denominador', 'Cumplimiento', 'Meta_Fijada', 'Meta_Nacional'])
            writer.writeheader()
            for r in reporte:
                writer.writerow({
                    'Centro': r['Centro'],
                    'Meta_ID': r['Meta_ID'],
                    'Indicador': r['Indicador'],
                    'Numerador': r['Numerador'],
                    'Denominador': r['Denominador'],
                    'Cumplimiento': r['Cumplimiento'],
                    'Meta_Fijada': r['Meta_Fijada'],
                    'Meta_Nacional': r['Meta_Nacional']
                })
        print(f"Reporte guardado en {output_path}")
    except Exception as e:
        print(e)

if __name__ == "__main__":
    calcular_meta_4()
