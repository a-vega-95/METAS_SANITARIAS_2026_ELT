import sys
import os
import csv
import openpyxl
import pyarrow.parquet as pq

# Add project root to path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(project_root, 'SRC'))

from modules.dataloaders import scan_rem_files, load_piv_data
from modules.utils import normalize_path
from config import (
    DIR_SERIE_P_ACTUAL, 
    PIV_FILE, 
    PREVALENCIA_HTA_15_24, 
    PREVALENCIA_HTA_25_44, 
    PREVALENCIA_HTA_45_64, 
    PREVALENCIA_HTA_65_MAS,
    AGNO_ACTUAL
)

def calcular_meta_5():
    print("=== Calculando Meta 5: Hipertensión Arterial (HTA) ===")
    
    # Meta 5: Cobertura Efectiva HTA (P4 Sección B)
    # Num: C34 + C35 (Personas 15-79 <140/90 + 80+ <150/90)
    SHEET = "P4"
    CELLS = ["C34", "C35"]
    
    try:
        mapping = scan_rem_files(DIR_SERIE_P_ACTUAL)
        piv_data = load_piv_data(PIV_FILE)
    except Exception as e:
        print(e)
        return

    # 1. Denominadores Estimados (PIV Estratificado)
    # Res. Exenta 650:
    # 15-24: 0.7%
    # 25-44: 10.6%
    # 45-64: 45.1%
    # 65+:   73.3%
    
    denominadores = {}
    
    for row in piv_data:
        centro = row.get('COD_CENTRO', '')
        edad = row.get('EDAD_EN_FECHA_CORTE')
        if edad is None: edad = -1
        estado = row.get('ACEPTADO_RECHAZADO', '')
        
        if estado == 'ACEPTADO':
            if centro not in denominadores:
                denominadores[centro] = 0
            
            factor = 0.0
            if 15 <= edad <= 24:
                factor = PREVALENCIA_HTA_15_24
            elif 25 <= edad <= 44:
                factor = PREVALENCIA_HTA_25_44
            elif 45 <= edad <= 64:
                factor = PREVALENCIA_HTA_45_64
            elif edad >= 65:
                factor = PREVALENCIA_HTA_65_MAS
                
            if factor > 0:
                denominadores[centro] += (1 * factor)
                
    # Redondear denominadores
    denominadores = {k: round(v) for k, v in denominadores.items()}
    
    # 2. Numeradores (REM)
    numeradores = {}
    
    for entry in mapping:
        # Normalize to base code
        raw_code = entry['code']
        # Try to strip trailing letters if numeric part exists
        real_code = raw_code
        if raw_code[-1].isalpha() and raw_code[:-1].isdigit():
             real_code = raw_code[:-1]
             
        file_path = entry['path']
        
        if real_code not in numeradores:
            numeradores[real_code] = 0
            
        if not os.path.exists(file_path): continue
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            if SHEET in wb.sheetnames:
                sheet = wb[SHEET]
                # Dynamic Search for C34+C35 equivalents
                # "PA < 140/90 mmHg" (usually Row 28/29)
                # "PA < 150/90 mmHg"
                
                rows_found = 0
                for row in sheet.iter_rows(min_row=1, max_row=100, values_only=True):
                    row_str = " ".join([str(c) for c in row[:5] if c])
                    
                    if "PA < 140/90" in row_str or "PA < 150/90" in row_str:
                        # Value usually in Col C (Index 2)
                        if len(row) > 2 and isinstance(row[2], (int, float)):
                            numeradores[real_code] += row[2]
                            rows_found += 1
                            
                    # Optimization: break if we found both?
                    # Careful if there are duplicates (e.g. by age group breakdown rows).
                    # But the "Total" rows are usually unique in the "Metas de Compensación" section.
                    # We continue scanning to be safe or break if we are sure.
                    # Given the dump, they appear sequentially.
                    
            wb.close()
        except: pass

    # Reporte
    all_centers = set(denominadores.keys()) | set(numeradores.keys())
    reporte = []
    
    total_num = 0
    total_den = 0
    
    for c in all_centers:
        num = numeradores.get(c, 0)
        den = denominadores.get(c, 0)
        cump = (num/den*100) if den > 0 else 0
        
        total_num += num
        total_den += den
        
        reporte.append({
            'Centro': c, 
            'Meta_ID': 'Meta 5',
            'Indicador': 'Cobertura HTA',
            'Numerador': num, 
            'Denominador': den, 
            'Cumplimiento_Actual': cump,
            'Meta_Fijada': 40.0,
            'Meta_Nacional': 45.0
        })
        
    print("\n=== RESULTADOS GLOBALES META 5 (HTA) ===")
    print(f"Numerador: {total_num}")
    print(f"Denominador (Est. por Factores): {total_den}")
    if total_den > 0:
        print(f"Cumplimiento: {total_num/total_den*100:.2f}%")
    print("Meta Fijada: 40.0%")
        
    # Guardar reporte
    output_path = normalize_path(r"DATOS\reporte_meta_5_preliminar.csv")
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Centro', 'Meta_ID', 'Indicador', 'Numerador', 'Denominador', 'Cumplimiento', 'Meta_Fijada', 'Meta_Nacional'])
            writer.writeheader()
            for r in reporte:
                # Map internal dict keys to CSV headers if needed
                writer.writerow({
                    'Centro': r['Centro'],
                    'Meta_ID': r['Meta_ID'],
                    'Indicador': r['Indicador'],
                    'Numerador': r['Numerador'],
                    'Denominador': r['Denominador'],
                    'Cumplimiento': r['Cumplimiento_Actual'],
                    'Meta_Fijada': r['Meta_Fijada'],
                    'Meta_Nacional': r['Meta_Nacional']
                })
        print(f"Reporte guardado en {output_path}")
    except Exception as e:
        print(f"Error escribiendo reporte: {e}")

if __name__ == "__main__":
    calcular_meta_5()
