import sys
import os
import csv
import openpyxl

# Add project root to path
current_dir = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(os.path.dirname(current_dir))
sys.path.append(os.path.join(project_root, 'SRC'))

from modules.dataloaders import scan_rem_files, load_piv_data
from config import DIR_SERIE_P_ACTUAL, PIV_FILE, PREVALENCIA_ASMA, PREVALENCIA_EPOC
from modules.utils import normalize_path

def to_num(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return val
    try:
        return float(val)
    except:
        return 0

def calcular_meta_7():
    print("=== Calculando Meta 7: Enfermedades Respiratorias (Asma/EPOC) ===")
    
    SHEET_TARGET = "P3"
    
    try:
        mapping = scan_rem_files(DIR_SERIE_P_ACTUAL)
        piv_data = load_piv_data(PIV_FILE)
    except Exception as e:
        print(e)
        return

    # 1. Denominadores Estimados (PIV)
    denominadores = {}
    
    for row in piv_data:
        centro = row.get('COD_CENTRO', '')
        edad = row.get('EDAD_EN_FECHA_CORTE')
        if edad is None: edad = -1
        estado = row.get('ACEPTADO_RECHAZADO', '')
        
        if estado == 'ACEPTADO':
            if centro not in denominadores:
                denominadores[centro] = 0
            
            # Asma 5+
            if edad >= 5:
                denominadores[centro] += (1 * PREVALENCIA_ASMA)
                
            # EPOC 40+
            if edad >= 40:
                denominadores[centro] += (1 * PREVALENCIA_EPOC)
                
    # Redondear
    denominadores = {k: round(v) for k, v in denominadores.items()}
    
    # 2. Numeradores (REM P3)
    numeradores = {}
    
    for entry in mapping:
        code = entry['code']
        file_path = entry['path']
        
        if code not in numeradores:
            numeradores[code] = 0
            
        if not os.path.exists(file_path): continue
        
        try:
             wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
             
             if SHEET_TARGET in wb.sheetnames:
                 ws = wb[SHEET_TARGET]
                 
                 for row in ws.iter_rows(min_row=1, max_row=300, values_only=True):
                     if not row or len(row) < 10: continue
                     
                     row_str = " ".join([str(c) for c in row[:5] if c])
                     
                     val_asma = 0
                     val_epoc = 0
                     
                     # ASMA
                     if "Asma" in row_str and "Controlado" in row_str:
                         # Total is at Index 2
                         total = to_num(row[2])
                         c5 = to_num(row[5]) # 0-4 Men
                         c6 = to_num(row[6]) # 0-4 Women
                         val_asma = total - (c5 + c6)
                         numeradores[code] += max(0, val_asma)
                         
                     # EPOC
                     if "EPOC" in row_str and "Control" in row_str and "Adecuado" in row_str:
                         # Sum from Index 21 onwards (40-44 years starts here? Verify based on P3 structure dump)
                         # Dump showed headers: 15-19 (Idx 3?), 20-24...
                         # Dump Row 10: '0 a 4 años', '5 a 9 años', ...
                         # Let's verify start index for 40+.
                         # 0-4: 5,6
                         # 5-9: 7,8
                         # 10-14: 9,10
                         # 15-19: 11,12
                         # 20-24: 13,14
                         # 25-29: 15,16
                         # 30-34: 17,18
                         # 35-39: 19,20
                         # 40-44: 21,22 -> CORRECT. Start summing from 21.
                         
                         current_sum = 0
                         for idx in range(21, len(row)):
                             current_sum += to_num(row[idx])
                         val_epoc = current_sum
                         numeradores[code] += val_epoc
                         
             wb.close()
        except: pass
            
    # Reporte
    reporte = []
    all_centers = set(denominadores.keys()) | set(numeradores.keys())
    
    total_num = 0
    total_den = 0
    
    for c in all_centers:
        num = numeradores.get(c, 0)
        den = denominadores.get(c, 0)
        cump = (num/den*100) if den > 0 else 0
        
        total_num += num
        total_den += den
        
        reporte.append({
            'Centro': c, 'Meta_ID': 'Meta 7', 'Indicador': 'Resp (Asma/EPOC)',
            'Numerador': num, 'Denominador': den, 'Cumplimiento': cump,
            'Meta_Fijada': 16.77, 'Meta_Nacional': 15.0
        })
        
    print("\n=== RESULTADOS GLOBALES META 7 ===")
    print(f"Numerador Total: {total_num}")
    print(f"Denominador Total (Est): {total_den}")
    if total_den > 0:
         print(f"Cumplimiento: {total_num/total_den*100:.2f}%")
         
    output_path = normalize_path(r"DATOS\reporte_meta_7_preliminar.csv")
    try:
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=['Centro', 'Meta_ID', 'Indicador', 'Numerador', 'Denominador', 'Cumplimiento', 'Meta_Fijada', 'Meta_Nacional'])
            writer.writeheader()
            for r in reporte:
                writer.writerow(r)
        print(f"Reporte guardado en {output_path}")
    except: pass
    
if __name__ == "__main__":
    calcular_meta_7()
