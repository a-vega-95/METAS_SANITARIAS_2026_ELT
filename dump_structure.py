import openpyxl
import os

A_FILE = r"DATOS\ENTRADA\SERIE_A\2025\DIC_2025\121305A.xlsm"
P_FILE = r"DATOS\ENTRADA\SERIE_P\2025\121305P.xlsm"

with open("debug_structure.txt", "w", encoding="utf-8") as f:
    # A03
    if os.path.exists(A_FILE):
        wb = openpyxl.load_workbook(A_FILE, read_only=True, data_only=True)
        if "A03" in wb.sheetnames:
            ws = wb["A03"]
            f.write("=== A03 SECTION D ===\n")
            for i, row in enumerate(ws.iter_rows(min_row=200, max_row=220, values_only=True), 200):
                line = f"Row {i}: {row}\n"
                f.write(line)
    
    # P4
    if os.path.exists(P_FILE):
        wb = openpyxl.load_workbook(P_FILE, read_only=True, data_only=True)
        sheet = "P4" if "P4" in wb.sheetnames else None
        if sheet:
            ws = wb[sheet]
            f.write("\n=== P4 ===\n")
            for i, row in enumerate(ws.iter_rows(min_row=1, max_row=300, values_only=True), 1):
                clean_row = [str(c) for c in row if c is not None]
                if clean_row:
                    line = f"Row {i}: {clean_row[:5]}\n" # Just first 5 cols usually have labels
                    f.write(line)
