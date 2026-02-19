import openpyxl
import os
import sys

# Adjust these paths if necessary based on list_dir results
A_FILE = r"DATOS\ENTRADA\SERIE_A\2025\DIC_2025\121305A.xlsm"
P_FILE = r"DATOS\ENTRADA\SERIE_P\2025\121305P.xlsm"

def inspect_a03():
    target_file = A_FILE
    if not os.path.exists(target_file):
        print(f"File not found: {target_file}")
        dir_path = os.path.dirname(target_file)
        if os.path.exists(dir_path):
            files = os.listdir(dir_path)
            if files:
                target_file = os.path.join(dir_path, files[0])
                print(f"Using alternate file: {target_file}")

    if not os.path.exists(target_file):
        print("No A03 file found.")
        return

    try:
        wb = openpyxl.load_workbook(target_file, read_only=True, data_only=True)
        if "A03" not in wb.sheetnames:
            print("Sheet A03 not found")
        else:
            ws = wb["A03"]
            print("\n--- Inspecting A03 Section D (Rows 204-216) ---")
            for i, row in enumerate(ws.iter_rows(min_row=204, max_row=216, values_only=True), 204):
                print(f"Row {i}: {row[:15]}")
    except Exception as e:
        print(f"Error A03: {e}")

def inspect_p04():
    print(f"Inspecting P04 in {P_FILE}")
    if not os.path.exists(P_FILE):
        print(f"File not found: {P_FILE}")
        return

    try:
        wb = openpyxl.load_workbook(P_FILE, read_only=True, data_only=True)
        print(f"Sheets in P file: {wb.sheetnames}")
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['NOMBRE', 'Control', 'MACROS']: continue
            
            print(f"\n--- Inspecting {sheet_name} ---")
            try:
                ws = wb[sheet_name]
                for i, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=True), 1):
                    content = " ".join([str(c) for c in row[:10] if c is not None])
                    if "Respiratoria" in content or "Asma" in content or "EPOC" in content:
                        print(f"Found keyword in {sheet_name} Row {i}: {row[:10]}")
            except:
                print(f"Could not read {sheet_name}")

    except Exception as e:
        print(f"Error P04: {e}")
    except Exception as e:
        print(f"Error P04: {e}")

if __name__ == "__main__":
    inspect_a03()
    inspect_p04()
